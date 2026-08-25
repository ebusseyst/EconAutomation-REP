import logging
from pathlib import Path
from typing import Any

import openpyxl as opxl
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)


def create_case_variables_excel_sheet(
    case_profile: Any,
    case_var_wb_path: Path,
    claimant_dir: Path,
) -> None:
    """
    Creates a new Excel workbook with the same structure as the template workbook.
    Populates the one worksheet with data from the case_profile dataclass.
    """

    # 1. Load the template workbook
    workbook = opxl.load_workbook(str(case_var_wb_path))

    # 2. Get the one worksheet
    sheet = workbook["REPORT_OUTPUTS"]

    # 3. Populate the worksheet with data from the case_profile dataclass.
    # The new layout has 4 side-by-side semantic groups:
    #   CLAIMANT INFO  → name col A (1), value col B (2)
    #   CASE INFO      → name col E (5), value col F (6)
    #   ATTORNEY INFO  → name col I (9), value col J (10)
    #   ECON INFO      → name col M (13), value col N (14)
    _COL_GROUPS = [(1, 2), (5, 6), (9, 10), (13, 14)]

    label_to_cell: dict[str, tuple[int, int]] = {}
    for name_col, value_col in _COL_GROUPS:
        for row_cells in sheet.iter_rows(min_row=3, min_col=name_col, max_col=name_col):
            c = row_cells[0]
            if c.value is not None and not isinstance(c, MergedCell):
                label_to_cell[str(c.value)] = (c.row, value_col)

    for field in case_profile.__dataclass_fields__.values():
        entry = label_to_cell.get(field.name)
        if entry is None:
            continue
        row, value_col = entry
        cell = sheet.cell(row=row, column=value_col)
        if isinstance(cell, MergedCell):
            for merge_range in sheet.merged_cells.ranges:
                if (
                    merge_range.min_row <= row <= merge_range.max_row
                    and merge_range.min_col <= value_col <= merge_range.max_col
                ):
                    top_left = sheet.cell(
                        row=merge_range.min_row, column=merge_range.min_col
                    )
                    if not isinstance(top_left, MergedCell):
                        top_left.value = getattr(case_profile, field.name)
                    break
        else:
            cell.value = getattr(case_profile, field.name)

    # 4. Save the workbook to the claimant's directory
    case_var_new_name = f"{case_profile.claimant_name_last}{case_profile.claimant_name_first_initial} - Case_Variables.xlsx"
    output_filepath = claimant_dir / case_var_new_name

    if output_filepath.exists():
        logger.warning("Skipping Case Variables: %s already exists.", output_filepath)
        return

    workbook.save(str(output_filepath))
