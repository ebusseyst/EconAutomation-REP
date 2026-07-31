import datetime as dt
import logging
import locale
import time
from decimal import Decimal
from pathlib import Path
from typing import Any, Optional
import platform

from pydantic import create_model, ConfigDict

import openpyxl as opxl
from openpyxl.cell.cell import MergedCell
import xlwings as xw

# Module's logger instance
logger = logging.getLogger(__name__)

# Maps REPORT_OUTPUTS variable_type column values → DataFormatterCore list keys.
# "Text" and "Integer" are intentionally absent: Text needs no reformatting,
# Integer is handled separately via reprocess_integers.
_VARIABLE_TYPE_MAP: dict[str, str] = {
    "Currency": "currency_values",
    "Date (Short)": "short_form_dates",
    "Date (Long)": "long_form_dates",
    "Percentage": "percentages",
    "Decimal/Float": "floats",
    "Integer": "integers",
}


class WorkbookInfoCore:
    """
    Core class for workbook information.
    """

    def __init__(
        self,
        workbook_filepath: Path,
        workbook_outputs_sheet_name: str = "REPORT_OUTPUTS",
    ) -> None:

        # DEFINING CLASS ATTRIBUTES
        self.workbook_path = workbook_filepath
        self.workbook_pathstr = str(workbook_filepath)
        self.workbook_name = Path(workbook_filepath).name.split(".")[0]
        self.workbook_outputs_sheet_name = workbook_outputs_sheet_name
        self.active_workbook = opxl.load_workbook(
            workbook_filepath, read_only=True, data_only=True
        )

        # DEFINING WORKBOOK VARIABLES
        self.workbook_variables_dict = self.define_workbook_variables_dict(
            workbook_name=self.workbook_name,
            active_workbook=self.active_workbook,
            workbook_outputs_sheet_name=self.workbook_outputs_sheet_name,
        )

        # DEFINING WORKBOOK CHARTS
        self.workbook_charts_dict = self.define_workbook_charts(
            workbook_name=self.workbook_name,
            active_workbook=self.active_workbook,
            workbook_outputs_sheet_name=self.workbook_outputs_sheet_name,
        )

        # DEFINING WORKBOOK TABLES
        self.workbook_tables_dict = self.define_workbook_tables(
            workbook_name=self.workbook_name,
            active_workbook=self.active_workbook,
            workbook_outputs_sheet_name=self.workbook_outputs_sheet_name,
        )

        # CLOSING THE WORKBOOK
        self.active_workbook.close()

    def define_workbook_variables_dict(
        self,
        workbook_name: str,
        active_workbook: opxl.Workbook,
        workbook_outputs_sheet_name: str,
    ) -> dict[str, tuple[str, str, str]]:
        """
        Defines the target cells for the instanced workbook. Data is pulled from dedicated "REPORT_OUTPUTS" tab on each
        relevant worksheet and returned as a nested dictionary (e.g., {"variable_name": ("worksheet_name", "cell_address")}).
        """
        try:
            workbook_variables_dict = {}
            workbook_outputs_sheet = active_workbook[workbook_outputs_sheet_name]
            if "Case" in workbook_name and "Variable" in workbook_name:
                # New layout: 4 side-by-side semantic groups.
                # Each group: (name_col, value_col, type_col) — 1-indexed.
                #   CLAIMANT INFO  A(1)  B(2)  C(3)
                #   CASE INFO      E(5)  F(6)  G(7)
                #   ATTORNEY INFO  I(9)  J(10) K(11)
                #   ECON INFO      M(13) N(14) O(15)
                _COL_GROUPS = [(1, 2, 3), (5, 6, 7), (9, 10, 11), (13, 14, 15)]
                for name_col, _value_col, type_col in _COL_GROUPS:
                    for row in workbook_outputs_sheet.iter_rows(
                        min_row=3,
                        min_col=name_col,
                        max_col=type_col,
                        max_row=110,
                        values_only=False,
                    ):
                        name_cell, value_cell, type_cell = row[0], row[1], row[2]
                        if isinstance(name_cell, MergedCell):
                            continue
                        variable_name = name_cell.value
                        variable_value = value_cell.value
                        variable_type = type_cell.value if type_cell.value else "Text"
                        if variable_name == "" or variable_name is None:
                            continue
                        if variable_value == "" or variable_value is None:
                            continue
                        workbook_variables_dict[str(variable_name)] = (
                            workbook_outputs_sheet_name,
                            f"{value_cell.coordinate}",
                            str(variable_type),
                        )
            for row in workbook_outputs_sheet.iter_rows(
                min_row=3, min_col=1, max_col=4, max_row=150, values_only=True
            ):
                variable_name = row[0]
                worksheet_name = row[1]
                cell_address = row[2]
                variable_type = row[3]
                if (
                    (variable_name == "" or variable_name is None)
                    or (worksheet_name == "" or worksheet_name is None)
                    or (cell_address == "" or cell_address is None)
                    or (variable_type == "" or variable_type is None)
                ):
                    logger.debug(
                        "DataExtractorCore.define_workbook_variables_dict: Empty column(s) found in REPORT_OUTPUTS tab."
                    )
                    continue
                workbook_variables_dict[f"{variable_name}"] = (
                    f"{worksheet_name}",
                    f"{cell_address}",
                    f"{variable_type}",
                )
            logger.info(
                f"DataExtractorCore.define_workbook_variables_dict: {workbook_name} variables dict: {workbook_variables_dict}"
            )
            return workbook_variables_dict
        except KeyError:
            logger.exception(
                f"DataExtractorCore.define_workbook_variables_dict: KeyError defining workbook variables for {workbook_name}"
            )
            raise
        except IndexError:
            logger.exception(
                f"DataExtractorCore.define_workbook_variables_dict: IndexError defining workbook variables for {workbook_name}"
            )
            raise
        except TypeError:
            logger.exception(
                f"DataExtractorCore.define_workbook_variables_dict: TypeError defining workbook variables for {workbook_name}"
            )
            raise

    def define_workbook_charts(
        self,
        workbook_name: str,
        active_workbook: opxl.Workbook,
        workbook_outputs_sheet_name: str,
    ) -> dict[str, str]:
        """
        Defines the worksheets with charts for the relevant workbook.
        """
        # Case_Variables uses a 4-column-group layout where columns J-K (10-11)
        # hold attorney values — the same columns this function reads for chart
        # config.  That workbook has no charts to extract, so bail early.
        if "Case" in workbook_name and "Variable" in workbook_name:
            return {}
        try:
            workbook_charts_dict = {}
            wb_outputs_sheet = active_workbook[workbook_outputs_sheet_name]
            for row in wb_outputs_sheet.iter_rows(
                min_row=3, min_col=10, max_col=11, max_row=40, values_only=True
            ):
                worksheet_name = row[0]
                chart_name = row[1]
                if worksheet_name == "" or worksheet_name is None:
                    logger.debug(
                        "DataExtractorCore.define_workbook_charts: Empty column(s) found in REPORT_OUTPUTS tab."
                    )
                    continue
                workbook_charts_dict[str(worksheet_name)] = str(chart_name)
            logger.info(
                f"DataExtractorCore.define_workbook_charts: {workbook_name} charts list: {workbook_charts_dict}"
            )
            return workbook_charts_dict
        except KeyError:
            logger.exception(
                f"DataExtractorCore.define_workbook_charts: KeyError defining workbook charts for {workbook_name}"
            )
            raise
        except IndexError:
            logger.exception(
                f"DataExtractorCore.define_workbook_charts: IndexError defining workbook charts for {workbook_name}"
            )
            raise
        except TypeError:
            logger.exception(
                f"DataExtractorCore.define_workbook_charts: TypeError defining workbook charts for {workbook_name}"
            )
            raise

    def define_workbook_tables(
        self,
        workbook_name: str,
        active_workbook: opxl.Workbook,
        workbook_outputs_sheet_name: str,
    ) -> dict[str, str]:
        """
        Defines the worksheets with tables for the relevant workbook.
        """
        # Same column-layout collision as define_workbook_charts — skip entirely.
        if "Case" in workbook_name and "Variable" in workbook_name:
            return {}
        try:
            workbook_tables_dict = {}
            wb_outputs_sheet = active_workbook[workbook_outputs_sheet_name]
            for row in wb_outputs_sheet.iter_rows(
                min_row=3, min_col=7, max_col=8, max_row=40, values_only=True
            ):
                worksheet_name = row[0]
                table_name = row[1]
                if (
                    worksheet_name == ""
                    or worksheet_name is None
                    or table_name == ""
                    or table_name is None
                ):
                    logger.debug(
                        "DataExtractorCore.define_workbook_tables: Empty column(s) found in REPORT_OUTPUTS tab."
                    )
                    continue
                workbook_tables_dict[str(worksheet_name)] = str(table_name)
            logger.info(
                f"DataExtractorCore.define_workbook_tables: {workbook_name} tables list: {workbook_tables_dict}"
            )
            return workbook_tables_dict
        except KeyError:
            logger.exception(
                f"DataExtractorCore.define_workbook_tables: KeyError defining workbook tables for {workbook_name}"
            )
            raise
        except IndexError:
            logger.exception(
                f"DataExtractorCore.define_workbook_tables: IndexError defining workbook tables for {workbook_name}"
            )
            raise
        except TypeError:
            logger.exception(
                f"DataExtractorCore.define_workbook_tables: TypeError defining workbook tables for {workbook_name}"
            )
            raise


class DataFormatterCore:
    """
    Formats relevant values extracted from the workbook.
    """

    def __init__(
        self,
        workbook_dataclass: Any,
        reformatting_lists_dict: dict[str, list[str]],
    ):
        # DEFINING CLASS ATTRIBUTES
        self.workbook_dataclass = workbook_dataclass
        self.reformatting_lists_dict = reformatting_lists_dict

        # CALLING MAIN REPROCESSING METHOD
        self.reprocess_all_relevant_data()

    # ── Private parsing helpers ──────────────────────────────────────────

    @staticmethod
    def _parse_date_value(date_value: str | dt.date) -> dt.date:
        """Converts a date value to a date object."""
        if isinstance(date_value, dt.date):
            return date_value
        else:
            formats = [
                "%Y-%m-%d",
                "%Y/%m/%d",
                "%m-%d-%Y",
                "%m/%d/%Y",
                "%m-%d-%y",
                "%m/%d/%y",
                "%B %d, %Y",
                "%b %d, %Y",
            ]
            for fmt in formats:
                try:
                    return dt.datetime.strptime(date_value, fmt).date()
                except ValueError:
                    continue
            raise ValueError(f"Unable to parse date value: {date_value}")

    @staticmethod
    def _parse_currency_value(currency_value: str | float | int) -> float:
        """
        Parses a currency value to ensure it is in the correct format.
        """
        if isinstance(currency_value, (int, float)):
            return float(currency_value)
        else:
            cleaned = currency_value.replace("$", "").replace(",", "").strip()
            return float(cleaned)

    @staticmethod
    def _parse_percentage_value(percentage_value: str | float) -> float:
        """
        Parses a percentage value to ensure it is in the correct format.
        """
        if isinstance(percentage_value, float):
            return percentage_value
        else:
            cleaned = (
                percentage_value.replace("$", "")
                .replace(",", "")
                .replace("%", "")
                .strip()
            )
            return float(cleaned)

    @staticmethod
    def _parse_float_value(float_value: str | float) -> float:
        """
        Parses a float value to ensure it is in the correct format.
        """
        if isinstance(float_value, float):
            return float_value
        else:
            cleaned = (
                float_value.replace("$", "").replace(",", "").replace("%", "").strip()
            )
            return float(cleaned)

    # ── Main reformatting methods ───────────────────────────────────────

    def reprocess_all_relevant_data(self) -> None:
        """
        Reprocesses all relevant data in the instanced extracted data dictionary.
        """
        for list_name, list_of_variables in self.reformatting_lists_dict.items():
            if list_name == "short_form_dates":
                self.reprocess_short_form_dates(short_form_dates_list=list_of_variables)
            elif list_name == "long_form_dates":
                self.reprocess_long_form_dates(long_form_dates_list=list_of_variables)
            elif list_name == "currency_values":
                self.reprocess_currency_values(currency_values_list=list_of_variables)
            elif list_name == "percentages":
                self.reprocess_percentages(percentages_list=list_of_variables)
            elif list_name == "floats":
                self.reprocess_floats(floats_list=list_of_variables)
            elif list_name == "integers":
                self.reprocess_integers(integers_list=list_of_variables)

    def reprocess_currency_values(self, currency_values_list: list[str]) -> None:
        """
        Reprocesses currency values to ensure they are in the correct format.
        """
        # Set locale for currency formatting
        try:
            locale.setlocale(locale.LC_ALL, "en_US.UTF-8")
        except locale.Error:
            locale.setlocale(locale.LC_ALL, "English_United States.1252")

        def format_currency(money_value: float) -> str:
            """
            Formats a float value as currency.
            """
            return locale.format_string("$%.0f", money_value, grouping=True)

        try:
            for value_name in currency_values_list:
                val = getattr(self.workbook_dataclass, value_name, None)
                if val is None or val == "":
                    continue
                setattr(
                    self.workbook_dataclass,
                    value_name,
                    format_currency(self._parse_currency_value(val)),
                )
            logger.debug(
                f"DataExtractorCore.reprocess_currency_values: Reprocessed currency values: {self.workbook_dataclass}"
            )
        except Exception as e:
            logger.exception(
                f"DataExtractorCore.reprocess_currency_values: Error reprocessing currency values: {e}"
            )

    def reprocess_short_form_dates(self, short_form_dates_list: list[str]) -> None:
        """
        Reprocesses provided date values to ensure they are in short-form format.
        """
        try:
            for date_value_name in short_form_dates_list:
                if getattr(self.workbook_dataclass, date_value_name, None) is None:
                    continue
                setattr(
                    self.workbook_dataclass,
                    date_value_name,
                    self._parse_date_value(
                        getattr(self.workbook_dataclass, date_value_name)
                    ).strftime("%m/%d/%Y"),
                )
            logger.debug(
                f"DataExtractorCore.reprocess_short_form_dates: Reprocessed short form dates: {self.workbook_dataclass}"
            )
        except Exception as e:
            logger.exception(
                f"DataExtractorCore.reprocess_short_form_dates: Error reprocessing short form dates: {e}"
            )

    def reprocess_long_form_dates(self, long_form_dates_list: list[str]) -> None:
        """
        Reprocesses provided date values to ensure they are in long-form format.
        """
        try:
            for date_value_name in long_form_dates_list:
                if getattr(self.workbook_dataclass, date_value_name, None) is None:
                    continue
                setattr(
                    self.workbook_dataclass,
                    date_value_name,
                    self._parse_date_value(
                        getattr(self.workbook_dataclass, date_value_name)
                    ).strftime("%B %d, %Y"),
                )
            logger.debug(
                f"DataExtractorCore.reprocess_long_form_dates: Reprocessed long form dates: {self.workbook_dataclass}"
            )
        except Exception as e:
            logger.exception(
                f"DataExtractorCore.reprocess_long_form_dates: Error reprocessing long form dates: {e}"
            )

    def reprocess_percentages(self, percentages_list: list[str]) -> None:
        """
        Reprocesses provided percentage values to ensure they are in the correct format.
        """
        try:
            for percentage_variable in percentages_list:
                if (
                    getattr(self.workbook_dataclass, percentage_variable, None) == ""
                    or getattr(self.workbook_dataclass, percentage_variable, None)
                    is None
                ):
                    continue
                parsed = (
                    self._parse_percentage_value(
                        getattr(self.workbook_dataclass, percentage_variable)
                    )
                    * 100
                )
                setattr(
                    self.workbook_dataclass,
                    percentage_variable,
                    f"{parsed:.2f}%",
                )
            logger.debug(
                f"DataExtractorCore.reprocess_percentages - reprocessed percentages: {self.workbook_dataclass}"
            )
        except Exception as e:
            logger.exception(
                f"DataExtractorCore.reprocess_percentages - Error reprocessing percentages: {e}"
            )

    def reprocess_floats(self, floats_list: list[str]) -> None:
        """
        Reprocesses provided float values to ensure they are in the correct format.
        """
        try:
            for float_variable in floats_list:
                if (
                    getattr(self.workbook_dataclass, float_variable, None) == ""
                    or getattr(self.workbook_dataclass, float_variable, None) is None
                ):
                    continue
                quantized = Decimal(
                    self._parse_float_value(
                        getattr(self.workbook_dataclass, float_variable)
                    )
                ).quantize(Decimal("0.00"))
                setattr(
                    self.workbook_dataclass,
                    float_variable,
                    str(quantized),  # TEMP: Convert to string for consistency
                )
            logger.debug(
                f"DataExtractorCore.reprocess_floats: Reprocessed floats: {self.workbook_dataclass}"
            )
        except Exception as e:
            logger.exception(
                f"DataExtractorCore.reprocess_floats: Error reprocessing floats: {e}"
            )

    def reprocess_integers(self, integers_list: list[str]) -> None:
        """
        Reprocesses provided values as integers (truncates any decimal portion).
        """
        try:
            for var_name in integers_list:
                val = getattr(self.workbook_dataclass, var_name, None)
                if val is None or val == "":
                    continue
                setattr(self.workbook_dataclass, var_name, int(float(val)))
            logger.debug(
                f"DataFormatterCore.reprocess_integers: Reprocessed integers: {self.workbook_dataclass}"
            )
        except Exception as e:
            logger.exception(
                f"DataFormatterCore.reprocess_integers: Error reprocessing integers: {e}"
            )


class DataExtractorCore:
    """
    Core class for extracting data from an active Excel workbook.
    """

    def __init__(
        self,
        workbook_pathstr: str,
        workbook_variables_dict: dict[str, tuple[str, ...]],
        workbook_charts_dict: dict[str, str],
        workbook_tables_dict: dict[str, str],
        temp_dir_path: Path,
    ):
        self.workbook_pathstr = workbook_pathstr
        self.workbook_variables_dict = workbook_variables_dict
        self.workbook_name = Path(workbook_pathstr).name.split(".")[0]
        self.temp_dir_path = temp_dir_path

        workbook_model_class = self.build_workbook_dataclass(
            self.workbook_name, self.workbook_variables_dict
        )

        # CALLING EXTRACTION METHODS
        self.workbook_dataclass = self.extract_data(
            self.workbook_pathstr, self.workbook_variables_dict, workbook_model_class
        )
        DataFormatterCore(
            workbook_dataclass=self.workbook_dataclass,
            reformatting_lists_dict=self.build_reformatting_lists_dict(
                self.workbook_variables_dict
            ),
        )
        self.extracted_image_paths: dict[str, Path] = {}
        self.extracted_image_paths.update(
            self.extract_charts(
                workbook_pathstr=self.workbook_pathstr,
                workbook_name=self.workbook_name,
                workbook_charts_dict=workbook_charts_dict,
                temp_dir_path=self.temp_dir_path,
            )
        )
        self.extracted_image_paths.update(
            self.extract_tables(
                workbook_pathstr=self.workbook_pathstr,
                workbook_name=self.workbook_name,
                workbook_tables_dict=workbook_tables_dict,
                temp_dir_path=self.temp_dir_path,
            )
        )

    def extract_data(
        self,
        workbook_pathstr: str,
        workbook_variables_dict: dict[str, tuple[str, ...]],
        workbook_dataclass: type[Any],
    ) -> Any:
        """
        Extracts data from the specified worksheet and returns a validated Pydantic model instance.
        """
        try:
            active_workbook = opxl.load_workbook(
                workbook_pathstr, read_only=True, data_only=True
            )
            data_dict = {}
            for variable_name, value_tuple in workbook_variables_dict.items():
                worksheet_name, cell_address = value_tuple[0], value_tuple[1]
                data_dict[variable_name] = active_workbook[worksheet_name][
                    cell_address
                ].value
            active_workbook.close()
            model_instance = workbook_dataclass(**data_dict)
            logger.info(
                f"DataExtractorCore.extract_data: Extracted data variables: {list(data_dict.keys())}"
            )
            logger.info(
                f"DataExtractorCore.extract_data: Created model instance: {model_instance}"
            )
            return model_instance
        except Exception as e:
            logger.exception(
                f"DataExtractorCore.extract_data: Error extracting data: {e}"
            )
            raise

    def build_workbook_dataclass(
        self,
        workbook_name: str,
        workbook_variables_dict: dict[str, tuple[str, ...]],
    ) -> type[Any]:
        """
        Returns a dynamically created Pydantic model class for the workbook's data fields.
        All fields are Optional[Any] with a None default so they can be populated after construction.
        validate_assignment=True triggers field validation when DataFormatterCore sets formatted values.
        """
        field_definitions: dict[str, Any] = {
            key: (Optional[Any], None) for key in workbook_variables_dict.keys()
        }
        return create_model(
            f"{workbook_name}Data",
            __config__=ConfigDict(
                validate_assignment=True, arbitrary_types_allowed=True
            ),
            **field_definitions,
        )

    @staticmethod
    def build_reformatting_lists_dict(
        workbook_variables_dict: dict[str, tuple],
    ) -> dict[str, list[str]]:
        """
        Builds a reformatting_lists_dict from the variable_type column (3rd element
        of each tuple) in workbook_variables_dict. Variables whose type is not in
        _VARIABLE_TYPE_MAP (e.g. "Text") or whose tuple has no type element are skipped.
        """
        result: dict[str, list[str]] = {}
        for var_name, value_tuple in workbook_variables_dict.items():
            if len(value_tuple) < 3:
                continue
            list_key = _VARIABLE_TYPE_MAP.get(value_tuple[2])
            if list_key is None:
                continue
            result.setdefault(list_key, []).append(var_name)
        return result

    def open_xw_workbook(self, workbook_pathstr: str) -> tuple[xw.App, xw.Book]:
        """
        Opens an Excel workbook using xlwings for chart/table extraction.
        """
        app = xw.App(visible=False)
        # Disable events so Workbook_Open / Worksheet_Activate handlers don't fire.
        app.api.EnableEvents = False
        # Force-disable ALL macros (msoAutomationSecurityForceDisable = 3).
        # EnableEvents only suppresses event handlers; Auto_Open subroutines bypass
        # it and can disconnect the COM object before we finish extraction.
        app.api.AutomationSecurity = 3
        wb = app.books.open(workbook_pathstr)
        time.sleep(0.5)
        return app, wb

    def extract_charts(
        self,
        workbook_pathstr: str,
        workbook_name: str,
        workbook_charts_dict: dict[str, str],
        temp_dir_path: Path,
    ) -> dict[str, Path]:
        """
        Extracts specified charts from the active workbook as png files.
        Returns a dict mapping chart_name → Path for each extracted image.
        """
        extracted: dict[str, Path] = {}
        try:
            app, wb = self.open_xw_workbook(workbook_pathstr)
            try:
                for sheet_name, chart_name in workbook_charts_dict.items():
                    try:
                        sheet = wb.sheets[sheet_name]
                    except Exception:
                        logger.warning(
                            f"DataExtractorCore.extract_charts: Sheet '{sheet_name}' "
                            f"not found in {workbook_name}, skipping chart '{chart_name}'."
                        )
                        continue
                    try:
                        charts_col = sheet.charts
                        has_chart = chart_name in charts_col
                        if not has_chart:
                            try:
                                available = [c.name for c in charts_col]
                            except Exception:
                                available = []
                            logger.warning(
                                f"DataExtractorCore.extract_charts: Chart '{chart_name}' not found "
                                f"in sheet '{sheet_name}' of {workbook_name}. "
                                f"Available chart names: {available}"
                            )
                            continue
                    except Exception:
                        logger.warning(
                            f"DataExtractorCore.extract_charts: Cannot access charts on "
                            f"sheet '{sheet_name}' in {workbook_name}, skipping '{chart_name}'."
                        )
                        continue
                    chart = sheet.charts[chart_name]
                    if platform.system() == "Darwin":
                        sheet.to_pdf(
                            temp_dir_path
                            / f"{workbook_name} - {sheet_name} - {chart_name}.pdf"
                        )
                    else:
                        out_path = (
                            temp_dir_path
                            / f"{workbook_name} - {sheet_name} - {chart_name}.png"
                        )
                        # Try direct file export first — no clipboard dependency.
                        # Chart.Export() infers PNG from the .png extension.
                        # If it silently produces EMF instead (detectable via header),
                        # fall back to the CopyPicture+clipboard approach.
                        _exported_valid = False
                        try:
                            chart.to_png(str(out_path))
                            with open(str(out_path), "rb") as _f:
                                _exported_valid = _f.read(4) == b"\x89PNG"
                        except Exception:
                            _exported_valid = False
                        if not _exported_valid:
                            from PIL import ImageGrab

                            xl_obj, _ = chart.api
                            for _retry in range(10):
                                try:
                                    xl_obj.CopyPicture(Appearance=1, Format=2)
                                    _im = ImageGrab.grabclipboard()
                                    if _im is None or isinstance(_im, list):
                                        raise AttributeError("clipboard empty")
                                    _im.save(str(out_path))
                                    break
                                except AttributeError:
                                    if _retry == 9:
                                        raise
                                    time.sleep(0.1)
                        extracted[chart_name] = out_path
                logger.info(
                    f"DataExtractorCore.extract_charts: Extracted charts from {workbook_name}"
                )
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
                try:
                    app.quit()
                except Exception:
                    pass

        except FileNotFoundError:
            logger.exception(
                f"DataExtractorCore.extract_charts: Error extracting charts: {workbook_pathstr} not found."
            )
            raise
        except Exception as e:
            logger.exception(
                f"DataExtractorCore.extract_charts: Error extracting charts: {e}"
            )
            raise
        return extracted

    def extract_tables(
        self,
        workbook_pathstr: str,
        workbook_name: str,
        workbook_tables_dict: dict[str, str],
        temp_dir_path: Path,
    ) -> dict[str, Path]:
        """
        Extracts tables from the specified worksheets as png files.
        Returns a dict mapping table_name → Path for each extracted image.
        """
        extracted: dict[str, Path] = {}
        try:
            app, wb = self.open_xw_workbook(workbook_pathstr)
            try:
                for sheet_name, table_name in workbook_tables_dict.items():
                    try:
                        sheet = wb.sheets[sheet_name]
                        if table_name not in sheet.tables:
                            logger.warning(
                                f"DataExtractorCore.extract_tables: Table '{table_name}' "
                                f"not found in sheet '{sheet_name}', skipping."
                            )
                            continue
                        table = sheet.tables[table_name]
                        out_path = (
                            temp_dir_path
                            / f"{workbook_name} - {sheet_name} - {table_name}.png"
                        )
                        from PIL import ImageGrab

                        rng = sheet.range(table.range.address)
                        for _retry in range(10):
                            try:
                                rng.api.CopyPicture(Appearance=1, Format=2)
                                _im = ImageGrab.grabclipboard()
                                if _im is None or isinstance(_im, list):
                                    raise AttributeError("clipboard empty")
                                _im.save(str(out_path))
                                break
                            except AttributeError:
                                if _retry == 9:
                                    raise
                                time.sleep(0.1)
                        extracted[table_name] = out_path
                    except Exception:
                        logger.warning(
                            f"DataExtractorCore.extract_tables: Could not extract "
                            f"table '{table_name}' from sheet '{sheet_name}' in "
                            f"{workbook_name}, skipping."
                        )
                        continue
                logger.info(
                    f"DataExtractorCore.extract_tables: Extracted tables from {workbook_name}"
                )
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
                try:
                    app.quit()
                except Exception:
                    pass
        except FileNotFoundError:
            logger.exception(
                f"DataExtractorCore.extract_tables: Error - {workbook_pathstr} not found."
            )
            raise
        except Exception as e:
            logger.exception(f"DataExtractorCore.extract_tables: Error - {e}")
            raise
        return extracted
