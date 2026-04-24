import logging
from pathlib import Path
import datetime
import locale
from decimal import Decimal
from typing import Any, ClassVar

import xlwings as xw
import openpyxl as opxl
import pandas as pd

# Module's logger instance
logger = logging.getLogger(__name__)

class WorkbookInfoCore:
    """
    Core class for workbook information.
    """
    def __init__(self, workbook_filepath: Path, workbook_outputs_sheet_name: str) -> :
        self.workbook = opxl.load_workbook(workbook_filepath)
        self.workbook_name = workbook_filepath.name
        self.workbook_outputs_sheet_name = workbook_outputs_sheet_name
        
        self.workbook_variables_dict = self.define_workbook_variables_dict(self.workbook, self.workbook_outputs_sheet_name)
        self
    
    def define_workbook_variables_dict(self, active_workbook:opxl.Workbook, workbook_outputs_sheet_name: str) -> dict[str, dict[str, Any]]:
        """
        Defines the target cells for the relevant workbook. Dictionaries are currently hardcoded and
        are individually defined for readability.
        """
        try:
            active_workbook_outputs_sheet = active_workbook[workbook_outputs_sheet_name]
            for col in active_workbook_outputs_sheet.iter_cols(min_row=3, min_col=1, max_col=4, max_row=100):
                values = [cell.value for cell in col]
                
            
        except Exception as e:
            logger.exception(f"DataExtractorCore.define_workbook_variables_dict: Error defining workbook variables: {e}")
            raise e
        
        return {}

class DataFormatterCore:
    """
    Formats relevant values extracted from the workbook.
    """
    def __init__(self, 
                workbook_name: str,
                extracted_data_dict: dict[str, Any], 
                short_form_dates_list: list[str] | None=None,
                long_form_dates_list: list[str] | None=None,
                currency_values_list: list[str] | None=None,
                percentages_list: list[str] | None=None,
                reformatted_floats_list: list[str] | None=None
                ):
        # DEFINING CLASS ATTRIBUTES
        self.workbook_name = workbook_name
        self.extracted_data_dict = extracted_data_dict
        
        if short_form_dates_list or long_form_dates_list:
            self.reprocess_dates(short_form_dates_list, long_form_dates_list)
        if currency_values_list:
            self.reprocess_currency_values(currency_values_list)
        if percentages_list:
            self.reprocess_percentages(percentages_list)
        if reformatted_floats_list:
            self.reprocess_floats(reformatted_floats_list)
            
        return self.extracted_data_dict

    def reprocess_currency_values(self, currency_values_list) -> None:
        """
        Reprocesses currency values to ensure they are in the correct format.
        """
        # Set locale for currency formatting
        locale.setlocale(locale.LC_ALL, "en_US.UTF-8")

        def format_currency(money_value: float) -> str:
            """
            Formats a float value as currency.
            """
            return locale.currency(money_value, symbol=True, grouping=True)
            
        try:
            for value_name in currency_values_list:
                if self.extracted_data_dict.get(value_name) == "" or value_name not in self.extracted_data_dict:
                    continue
                self.extracted_data_dict[value_name] = format_currency(self.extracted_data_dict[value_name])
            logger.info(f"DataExtractorCore.reprocess_currency_values: Reprocessed currency values: {self.extracted_data_dict}")
        except Exception as e:
            logger.exception(f"DataExtractorCore.reprocess_currency_values: Error reprocessing currency values: {e}")

            
    def reprocess_dates(self, short_form_dates_list: list[str] | None = None, long_form_dates_list: list[str] | None = None) -> None:
        """
        Reprocesses provided date values to ensure they are in the correct format.
        """
        if short_form_dates_list:
            try:
                for date_value_name in short_form_dates_list:
                    if self.extracted_data_dict.get(date_value_name) is None:
                        continue
                    self.extracted_data_dict[date_value_name] = self.extracted_data_dict[date_value_name].date()
                    self.extracted_data_dict[date_value_name] = datetime.datetime.strftime(self.extracted_data_dict[date_value_name], "%m/%d/%Y")
                    self.extracted_data_dict[date_value_name] = str(self.extracted_data_dict[date_value_name])
                logger.info(f"DataExtractorCore.reprocess_dates: Reprocessed short form dates: {self.extracted_data_dict}")
            except Exception as e:
                logger.exception(f"DataExtractorCore.reprocess_dates: Error reprocessing short form dates: {e}")
        
        if long_form_dates_list:
            try:
                for date_value_name in long_form_dates_list:
                    if self.extracted_data_dict.get(date_value_name) is None:
                        continue
                    self.extracted_data_dict[date_value_name] = self.extracted_data_dict[date_value_name].date()
                    self.extracted_data_dict[date_value_name] = datetime.datetime.strftime(self.extracted_data_dict[date_value_name], "%B %#d, %Y")
                    self.extracted_data_dict[date_value_name] = str(self.extracted_data_dict[date_value_name])
                logger.info(f"DataExtractorCore.reprocess_dates: Reprocessed long form dates: {self.extracted_data_dict}")
            except Exception as e:
                logger.exception(f"DataExtractorCore.reprocess_dates: Error reprocessing long form dates: {e}")
        
    def reprocess_percentages(self, percentages_list: list[str]) -> None:
        """
        Reprocesses provided percentage values to ensure they are in the correct format.
        """
        if percentages_list:
            try:
                for percentage_value_name in percentages_list:
                    if self.extracted_data_dict.get(percentage_value_name) == "" or self.extracted_data_dict.get(percentage_value_name) is None:
                        continue
                    self.extracted_data_dict[percentage_value_name] = self.extracted_data_dict[percentage_value_name] * 100
                    self.extracted_data_dict[percentage_value_name] = str(f"{self.extracted_data_dict[percentage_value_name]:.2f}%")
                logger.info(f"DataExtractorCore.reprocess_percentages: Reprocessed percentages: {self.extracted_data_dict}")
            except Exception as e:
                logger.exception(f"DataExtractorCore.reprocess_percentages: Error reprocessing percentages: {e}")
        
    def reprocess_floats(self, floats_list: list[str]) -> None:
        """
        Reprocesses provided float values to ensure they are in the correct format.
        """
        if floats_list:
            try:
                for float_value_name in floats_list:
                    if self.extracted_data_dict.get(float_value_name) == "" or self.extracted_data_dict.get(float_value_name) is None:
                        continue
                    self.extracted_data_dict[float_value_name] = Decimal(self.extracted_data_dict[float_value_name]).quantize(Decimal("0.00"))
                    self.extracted_data_dict[float_value_name] = str(self.extracted_data_dict[float_value_name]) # TEMP: Convert to string for consistency
                logger.info(f"DataExtractorCore.reprocess_floats: Reprocessed floats: {self.extracted_data_dict}")
            except Exception as e:
                logger.exception(f"DataExtractorCore.reprocess_floats: Error reprocessing floats: {e}")

class DataExtractorCore:
    """
    Core class for extracting data from an active Excel workbook.
    """
    def __init__(self, active_workbook_filepath: Path, data_formatter: DataFormatterCore):
        """
        Initializes the DataExtractorCore class.
        
        Args:
            active_workbook_filepath (Path): The filepath to the active workbook.
            data_formatter (DataFormatterCore): The data formatter object to use for reprocessing data.
        """
        self.data_formatter = data_formatter
        self.active_workbook = opxl.load_workbook(active_workbook_filepath, data_only=True) # Returns only computed values of formulas
    
    def load_workbook(self, active_workbook_path: Path):
        """
        Loads the active workbook.
        """
        self.active_workbook = opxl.load_workbook(active_workbook_path, data_only=True) # Returns only computed values of formulas
    
    def extract_data(self, workbook_targets_dict: dict[str, dict[str, str]]):
        """
        Extracts data from the specified worksheet based on its target cells subdictionaries.
        """
        extracted_dict = {}
        try:
            for worksheet_name, target_cells_dict in workbook_targets_dict.items():
                active_worksheet = self.active_workbook[worksheet_name]
                worksheet_targets_dict = workbook_targets_dict[worksheet_name]
                for value_name, cell_address in worksheet_targets_dict.items():
                    extracted_dict[value_name] = active_worksheet[cell_address].value
            logger.info(f"DataExtractorCore: Extracted data from {self.active_workbook}")
            logger.info(f"DataExtractorCore: Extracted data dict: {extracted_dict}")
            return extracted_dict
        except Exception as e:
            logger.exception(f"DataExtractorCore.extract_data: Error extracting data: {e}")
            return {}
    
    # def extract_tables(self, workbook_tables_list: list[str]=None):
    #     """
    #     Extracts tables from the specified worksheets based on their target table subdictionaries.
    #     """
    #     extracted_tables_dict = {}
    #     try:
    #         for worksheet_name in workbook_tables_list:
    #             for table_name, table in self.active_workbook[worksheet_name].tables.items():
    #                 df = pd.read_excel(self.active_workbook_path, sheet_name=worksheet_name, header=0, usecols=table.ref)
    #                 extracted_tables_dict[worksheet_name][table_name] = df
    #         logger.info(f"DataExtractorCore: Extracted tables from {self.active_workbook_path}")
    #         logger.info(f"DataExtractorCore: Extracted tables dict: {extracted_tables_dict}")
    #         return extracted_tables_dict
    #     except Exception as e:
    #         logger.exception(f"DataExtractorCore.extract_tables: Error extracting tables: {e}")
    #         return None
    
    